from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import sys




count=0

name=0
task_type=0
sheet=0
task_name=''
task_name_w_yml=''
sheet_name=''
name_line=''
temp_name_line=''
temp_line=''
start_rest_input=0
list_column=[]
temp_list=[]
temp_column=''
if_hit=0
file_new=open('column_names.yml', 'w')

overall_file1=open(sys.argv[1], 'r')
over_lines=overall_file1.readlines()

for overline in over_lines:
    # print(overline+'\n')
    temp_line='Task= '+overline.strip()+'\n'
    file_new.writelines(temp_line)
    file1 = open(overline.strip(), 'r')
    Lines = file1.readlines()

    count=0

    name=0
    task_type=0
    sheet=0
    task_name=''
    task_name_w_yml=''
    sheet_name=''
    name_line=''
    temp_name_line=''
    temp_line=''
    start_rest_input=0
    list_column=[]
    temp_list=[]
    temp_column=''
    if_hit=0
    for line in Lines:

        if(line.find('- name')!=-1):
            name=1
            name_line=line.rstrip()
            list_column=['number']
            temp_column=''
            # print('line0:'+line)
        else:
            if(name==1): #hit name line
                # if line.find('with_items')!=-1:
                #     print('task_type='+str(task_type))
                #     print('line='+line)
                #collect column titles
                if (line.find('item.')!=-1) and (line.find('#')==-1): #hit column
                    
                    for j in range (1, line.count('item.')+1):
                        temp_column=((line.strip()).split('item.')[j]).split('}')[0].strip()
                        #print("##################"+temp_column)
                        if_hit=0
                        for i in list_column:
                            if temp_column==i:
                                if_hit=1
                        if if_hit==0:
                            #print("!!!!!!!!!!!!!!!!!"+temp_column)
                            list_column.append(temp_column)
                        #print(list_column)
                if(line.find('hosts')!=-1): #not task
                    name=0            
                elif (line.find('aci_rest')!=-1): # task=aci_rest
                    task_type=1
                    temp_line=''
                    temp_line=temp_line+line.rstrip()+'\n'
                    # print('line1:'+line)
                    # task_name=task_name+'.yml'
                    # file_new=open(task_name, 'w')
                    # print(name_line)
                    # print(line.rstrip())
                elif (line.find('aci_')!=-1 and line.find('aci_login')==-1 and line.find('with_items')==-1): #task=aci_module
                    task_name=((line.strip()).split(':'))[0] #task_name=aci_module_name
                    task_type=2
                    temp_line=line.rstrip()+'\n'
                    temp_name_line=name_line+'\n'
                    # print(name_line)
                    print(line.rstrip())
                    # print('line2:'+line)
                    # task_name_w_yml='create_'+task_name+'.yml'
                    # file_new=open(task_name_w_yml, 'w')
                    # file_new.writelines('#It is generated by the script with the sourcing file as:'+sys.argv[1]+'\n')
                    # file_new.writelines(temp_name_line)
                    # file_new.writelines(temp_line)

                elif (task_name=='') and (task_type==1): #task=aci_rest
                    if(line.find('": {')!=-1): #find=target_name
                        temp_name_line=name_line+'\n'
                        start_rest_input=1
                        #task_name=name of the task for aci_rest call
                        task_name=line.split('"')[1]
                        task_name_w_yml='create_'+line.split('"')[1]+'.yml'
                        # file_new=open(task_name_w_yml, 'w')
                        # file_new.writelines('#It is generated by the script with the sourcing file as:'+sys.argv[1]+'\n')
                        # #first line about "- name"
                        # file_new.writelines(temp_name_line)
                        # #file_new=open(task_name, 'w')
                        # temp_line=temp_line+line.rstrip()+'\n'
                        # # print( temp_line)
                        # file_new.writelines(temp_line)
                    elif start_rest_input==0:
                        temp_line=temp_line+line.rstrip()+'\n'
                        # print(line.rstrip()) 
                    else:
                        print('!'+line.rstrip())
                        temp_line=line.rstrip()+'\n'
                        # file_new.writelines(temp_line.rstrip())
                                    
                elif(task_type!=0)and (line.find('with_items')!=-1): #find sheet name 
                    # print ('get into it!!!!!!\n')
                    sheet_name=(((line.split('{{'))[1]).split('}'))[0].strip()
                    temp_line=line.rstrip()+'\n'
                    if len(task_name)>29:
                        temp_line=(((line.split('{{'))[0]).rstrip("'")).rstrip('"')+"\"{{ "+task_name[0:29]+" }}\""+'\n'
                    else:
                        temp_line=(((line.split('{{'))[0]).rstrip("'")).rstrip('"')+"\"{{ "+task_name+" }}\""+'\n'
                    # file_new.writelines(temp_line)
                    
                    # print(line.rstrip())
                    #print(task_name+'  '+sheet_name)
                    # workbook_name=sheet_name+'.xlsx' #v1 workbook name=with_items name

                
                    # workbook_name=task_name+'.xlsx'
                    # wb = Workbook()
                    # ws1=wb.active
                    # ws1.title='validate_apic_id'
                    # ws1.append(['number','site','environment','function','factory'])


                    # file_new.writelines('#It is generated by the script with the sourcing file as:'+sys.argv[1]+'\n')
                    #first line about "- name"
                    # temp_name_line='validate_apic_id'+'\n'
                    # file_new.writelines(temp_name_line)
                    #file_new=open(task_name, 'w')
                    temp_list=['number','site','environment','function','factory']
                    temp_line=str(temp_list)+'\n'
                    # print( temp_line)
                    # file_new.writelines(temp_line)


                    file_new=open('column_names.yml', 'a')
                    #ws1=wb.create_sheet(title=sheet_name) #v1 sheet name=with_items name
                    if len(task_name)>29:
                        temp_line=task_name[0:29]
                    else:
                        temp_line=task_name
                    #add column titles
                    
                    temp_line='Sheet= '+temp_line.strip()+'\n'
                    file_new.writelines(temp_line)

                    

                    

                    # ws1.append(list_column)
                    temp_line='Columns= '+str(list_column)+'\n'
                    file_new.writelines(temp_line)
                    file_new.writelines('\n')

                    
                    # wb.save(filename=workbook_name)
                    #file_new.writelines('\n')

                elif(task_type!=0):
                    temp_line=line.rstrip()+'\n'
                    # file_new.writelines(temp_line)
                    # if(line.strip()==''):
                    #     #print(task_name+'  ')
                    #     name=0
                    #     task_type=0
                    #     task_name=''
                    #     task_name_w_yml=''
                    #     start_rest_input=0
                    #     file_new.writelines('# the end\n')
                    #     file_new.close()

                    #print('!!'+line.rstrip())
                # else:
                #     # print('line3:'+line)
                        


        # count += 1
        # print("Line{}: {}".format(count, line.strip()))